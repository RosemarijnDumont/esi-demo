import unittest
import pandas as pd
from test_data import generate_test_data
from export_script import export_to_excel
import openpyxl

class TestReportExport(unittest.TestCase):

    def setUp(self):
        self.test_df = generate_test_data()
        self.pre_fix_filename = "pre_fix_report.xlsx"
        self.post_fix_filename = "post_fix_report.xlsx"

    def _check_excel_formatting(self, filename):
        """
        Helper to check Excel file for basic formatting retention.
        This is a simplified check and may require manual verification.
        """
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active

            # Check Date_MM/DD/YYYY column (assuming it's the first column in the test data, B in excel)
            # We are checking for the presence of a format code that implies date formatting
            # This is a heuristic and might need to be more robust depending on the actual fix.
            date_cell_b2_number_format = sheet['B2'].number_format
            self.assertIn('m/d/yyyy', date_cell_b2_number_format.lower())

            # Check Date_DD-MMM-YY column (C in excel)
            date_cell_c2_number_format = sheet['C2'].number_format
            self.assertIn('dd-mmm-yy', date_cell_c2_number_format.lower())

            # Check Currency_USD column (D in excel)
            currency_cell_d2_number_format = sheet['D2'].number_format
            self.assertIn('$', currency_cell_d2_number_format.lower())
            self.assertIn('0.00', currency_cell_d2_number_format.lower())

            # Check Currency_EUR column (E in excel)
            currency_cell_e2_number_format = sheet['E2'].number_format
            self.assertIn('€', currency_cell_e2_number_format.lower())
            self.assertIn('0.00', currency_cell_e2_number_format.lower())

            print(f"\nBasic formatting check passed for {filename}.")
            return True
        except Exception as e:
            print(f"\nError during formatting check for {filename}: {e}")
            print("Manual verification of exported Excel file is recommended.")
            return False

    def test_pre_fix_export(self):
        print("\n--- Running Pre-Fix Export Test ---")
        export_to_excel(self.test_df, self.pre_fix_filename)
        print(f"Please manually review '{self.pre_fix_filename}' for lost formatting.")
        # In a real scenario, you'd likely have a more automated way to assert this failure
        # For now, we rely on manual review as per the task plan.


    def test_post_fix_export(self):
        print("\n--- Running Post-Fix Export Test ---")
        # Simulate the fix being applied to export_script.py before this test runs
        # For this test, we assume export_script.py has been modified to correctly format.
        
        # To properly test the fix, the `export_script.py` would need to be updated
        # to use a library like `xlsxwriter` or `openpyxl` with explicit formatting.
        # For demonstration, we'll assume the export_to_excel function is now 'fixed'.
        
        # A more complete fix would involve:
        # from openpyxl import Workbook
        # from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side
        # # ... inside export_to_excel ...
        # writer = pd.ExcelWriter(filename, engine='openpyxl')
        # dataframe.to_excel(writer, sheet_name='Report', index=False)
        # workbook = writer.book
        # sheet = writer.sheets['Report']
        #
        # # Apply styles (example for date and currency)
        # date_style = NamedStyle(name='datetime', number_format='MM/DD/YYYY')
        # currency_usd_style = NamedStyle(name='currency_usd', number_format='$#,##0.00')
        # currency_eur_style = NamedStyle(name='currency_eur', number_format='€#,##0.00')
        #
        # # Iterate and apply styles (example for specific columns)
        # for row_idx, row in enumerate(dataframe.iterrows()):
        #     # Apply date style to column 1 (Date_MM/DD/YYYY)
        #     sheet.cell(row=row_idx+2, column=2).style = date_style # +2 for header and 0-index
        #     # Apply date style to column 2 (Date_DD-MMM-YY)
        #     sheet.cell(row=row_idx+2, column=3).style = date_style # +2 for header and 0-index
        #     # Apply currency USD style to column 3 (Currency_USD)
        #     sheet.cell(row=row_idx+2, column=4).style = currency_usd_style
        #     # Apply currency EUR style to column 4 (Currency_EUR)
        #     sheet.cell(row=row_idx+2, column=5).style = currency_eur_style
        # writer.save()

        export_to_excel(self.test_df, self.post_fix_filename)
        self.assertTrue(self._check_excel_formatting(self.post_fix_filename),
                        "Post-fix Excel report did not retain expected formatting. Manual verification needed.")
        print(f"Please manually review '{self.post_fix_filename}' to confirm all acceptance criteria are met.")

    def test_no_regression_on_other_reports(self):
        print("\n--- Running Regression Test (Placeholder) ---")
        # This is a placeholder for actual regression tests.
        # In a real scenario, you would export other reports and verify their integrity.
        # For example, calling other export functions and checking their outputs.
        print("Performing a dummy regression check: ensuring test data generation is consistent.")
        another_df = generate_test_data()
        pd.testing.assert_frame_equal(self.test_df, another_df)
        print("Dummy regression check passed. Further manual regression testing is recommended.")

if __name__ == '__main__':
    unittest.main()